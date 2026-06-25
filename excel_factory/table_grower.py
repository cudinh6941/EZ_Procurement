"""
table_grower.py
────────────────
Logic co giãn 1 vùng "bảng item" trong sheet để khớp số lượng item thực tế,
dùng CHUNG cho cả PMT và CBE (thay vì 2 hàm viết tay riêng, gần như giống
nhau, như bản cũ).

Vùng item trong template có sẵn N0 "slot" (1 slot = 1 item, có thể chiếm
nhiều dòng vật lý do merge dọc — ví dụ CBE item đầu chiếm 2 dòng).
Khi số item thực tế (N) khác N0:
  - N < N0 : xoá nội dung (giữ style) các slot dư, không đụng tới layout.
  - N > N0 : chèn thêm (N - N0) dòng đơn (1 dòng/item) ngay trước vùng
             summary, copy style từ slot cuối cùng của template.
Trả về danh sách physical_row cho mỗi item (idx 0..N-1) + offset mà vùng
summary phía dưới đã bị dịch xuống.
"""
from __future__ import annotations

from copy import copy
from dataclasses import dataclass

from openpyxl.worksheet.worksheet import Worksheet
from .excel_schema import ItemTableSchema


@dataclass
class GrowResult:
    item_rows: list[int]      # physical row dùng để ghi item idx (mỗi item 1 dòng ghi)
    summary_offset: int       # số dòng đã chèn thêm phía trên vùng summary (>=0)


def _unmerge_in_range(ws: Worksheet, from_row: int, to_row: int):
    for rng in [str(r) for r in ws.merged_cells.ranges
                if r.min_row >= from_row and r.max_row <= to_row]:
        ws.unmerge_cells(rng)


def _copy_row_style(ws: Worksheet, source_row: int, target_row: int, max_col: int):
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def grow_item_table(ws: Worksheet, schema: ItemTableSchema, n_items: int) -> GrowResult:
    n_template = schema.n_template_items
    max_col = schema.max_col

    if n_items <= n_template:
        # Dùng lại các slot có sẵn theo đúng thứ tự; slot dư thì xoá nội dung,
        # GIỮ style/merge nguyên vẹn (không đụng layout khi không cần).
        item_rows = [r for r, _h in schema.template_item_slots[:max(n_items, 1)]]
        for r, _h in schema.template_item_slots[n_items:]:
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).value = None
        return GrowResult(item_rows=item_rows[:n_items], summary_offset=0)

    # n_items > n_template: cần chèn thêm dòng mới ngay trước summary_anchor_row
    extra = n_items - n_template
    insert_at = schema.summary_anchor_row

    _unmerge_in_range(ws, insert_at, insert_at + extra + 10)
    ws.insert_rows(insert_at, amount=extra)

    # style mẫu lấy từ slot cuối cùng của template (thường là dòng đơn, height=1)
    style_src_row, _h = schema.template_item_slots[-1]

    item_rows = [r for r, _h in schema.template_item_slots]  # n_template slot có sẵn
    for i in range(extra):
        new_row = insert_at + i
        _copy_row_style(ws, style_src_row, new_row, max_col)
        item_rows.append(new_row)

    return GrowResult(item_rows=item_rows, summary_offset=extra)
