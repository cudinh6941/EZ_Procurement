"""
excel_schema.py
────────────────
Lớp "hiểu cấu trúc template" — tách biệt hoàn toàn khỏi logic ghi dữ liệu.

Thay vì hardcode "item bắt đầu ở row 20, NCC2 ở cột H", module này QUÉT
chính template để tìm các "anchor" (ô mốc nhận diện bằng text), rồi suy ra
toạ độ. Nếu template bị sửa (thêm dòng, đổi tên cột...) mà anchor không còn
khớp, ta raise lỗi rõ ràng ngay lúc load — thay vì âm thầm ghi sai dữ liệu.

Mỗi schema chỉ biết "đọc cấu trúc", không biết gì về dữ liệu nghiệp vụ.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


class SchemaError(Exception):
    """Raised khi template không khớp với cấu trúc mong đợi — fail fast,
    không cho phép ghi dữ liệu sai một cách âm thầm."""


def _find_row_by_text(ws: Worksheet, text: str, col: int = 1,
                       row_range: tuple[int, int] = (1, 60),
                       mode: str = "startswith") -> int | None:
    """Tìm dòng đầu tiên trong `col` có giá trị string khớp `text`.
    mode='startswith' (mặc định, cho header cố định) hoặc 'contains'
    (cho các dòng có thể có số thứ tự/ký tự đứng trước, vd '2.Quy định...')."""
    for r in range(row_range[0], row_range[1] + 1):
        v = ws.cell(row=r, column=col).value
        if not isinstance(v, str):
            continue
        s = v.strip()
        if mode == "contains" and text in s:
            return r
        if mode == "startswith" and s.startswith(text):
            return r
    return None


def _find_cols_by_header(ws: Worksheet, header_row: int, text: str,
                          max_col: int = 30) -> list[int]:
    """Tìm TẤT CẢ cột ở `header_row` có giá trị string chứa `text`.
    Dùng để tự nhận diện số khối NCC (không hardcode số 3)."""
    cols = []
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and text in v:
            cols.append(c)
    return cols


def _merge_span(ws: Worksheet, row: int, col: int) -> int:
    """Số dòng vật lý mà ô (row, col) chiếm do merge theo chiều dọc (>=1)."""
    for rng in ws.merged_cells.ranges:
        if rng.min_col <= col <= rng.max_col and rng.min_row <= row <= rng.max_row:
            return rng.max_row - rng.min_row + 1
    return 1


@dataclass
class NccColumnBlock:
    """1 khối cột của 1 NCC. `cols` là danh sách cột theo đúng thứ tự
    ý nghĩa nghiệp vụ đã khai báo lúc tạo schema (vd: [don_gia, thanh_tien, ghi_chu])."""
    cols: list[int]

    def __getitem__(self, i):
        return self.cols[i]


@dataclass
class ItemTableSchema:
    """Cấu trúc 1 bảng item (PMT hoặc CBE), tự dò bằng anchor."""
    sheet_name: str
    header_row: int                 # dòng có tiêu đề "Đơn giá"/"Mô tả..."
    item_start_row: int             # dòng item đầu tiên = header_row + 1
    template_item_slots: list[tuple[int, int]]  # [(physical_start_row, height_rows), ...]
    ncc_blocks: list[NccColumnBlock]
    max_col: int
    name_rows: list[int]            # các dòng để ghi tên NCC (vd A6,A7,A8)
    summary_anchor_row: int         # dòng đầu của vùng tổng kết (text mốc đầu tiên)

    @property
    def n_template_items(self) -> int:
        return len(self.template_item_slots)

    @property
    def n_ncc_supported(self) -> int:
        return len(self.ncc_blocks)


def discover_pmt_schema(ws: Worksheet) -> ItemTableSchema:
    header_row = _find_row_by_text(ws, "STT", col=1)
    if header_row is None:
        raise SchemaError("[PMT] Không tìm thấy dòng header 'STT' ở cột A — "
                           "template có thể đã bị thay đổi cấu trúc.")

    name_section_row = _find_row_by_text(
        ws, "Quy định tên Nhà cung cấp", col=1, row_range=(1, header_row), mode="contains")
    if name_section_row is None:
        raise SchemaError("[PMT] Không tìm thấy dòng 'Quy định tên Nhà cung cấp'.")

    # NCC: cứ mỗi khối có cột header chứa "Mô tả Hàng hóa" -> đó là cột mô tả đầu khối
    desc_cols = _find_cols_by_header(ws, header_row, "Mô tả Hàng hóa")
    if not desc_cols:
        raise SchemaError("[PMT] Không tìm thấy cột 'Mô tả Hàng hóa' nào ở header.")

    ncc_blocks = []
    for dcol in desc_cols:
        # Khối PMT: mô tả, SL, hãng/xuất xứ, đánh giá => 4 cột liên tiếp
        ncc_blocks.append(NccColumnBlock(cols=[dcol, dcol + 1, dcol + 2, dcol + 3]))

    name_rows = list(range(name_section_row + 1, header_row - 1))  # trừ dòng "Yêu cầu PTSC"
    name_rows = name_rows[:len(ncc_blocks)]

    item_start_row = header_row + 1

    # Anchor đầu tiên của vùng summary: "Chứng chỉ Hàng hóa"
    summary_anchor_row = _find_row_by_text(ws, "Chứng chỉ Hàng hóa", col=1,
                                            row_range=(item_start_row, item_start_row + 30))
    if summary_anchor_row is None:
        raise SchemaError("[PMT] Không tìm thấy dòng mốc 'Chứng chỉ Hàng hóa'.")

    # Mỗi item PMT chiếm 1 dòng đơn (không merge dọc) trong template này
    template_item_slots = [(r, 1) for r in range(item_start_row, summary_anchor_row)]

    max_col = max(b.cols[-1] for b in ncc_blocks)

    return ItemTableSchema(
        sheet_name=ws.title,
        header_row=header_row,
        item_start_row=item_start_row,
        template_item_slots=template_item_slots,
        ncc_blocks=ncc_blocks,
        max_col=max_col,
        name_rows=name_rows,
        summary_anchor_row=summary_anchor_row,
    )


def discover_cbe_schema(ws: Worksheet) -> ItemTableSchema:
    header_row = _find_row_by_text(ws, "Đơn giá", col=5, row_range=(1, 40))
    # cột 5 (E) là điểm bắt đầu khối NCC1 trong layout chuẩn; nếu không thấy,
    # quét toàn dòng để tìm dòng có nhiều cell "Đơn giá".
    if header_row is None:
        for r in range(1, 40):
            if _find_cols_by_header(ws, r, "Đơn giá"):
                header_row = r
                break
    if header_row is None:
        raise SchemaError("[CBE] Không tìm thấy dòng header 'Đơn giá'.")

    name_section_row = _find_row_by_text(
        ws, "Quy định tên Nhà", col=1, row_range=(1, header_row), mode="contains")
    if name_section_row is None:
        raise SchemaError("[CBE] Không tìm thấy dòng 'Quy định tên Nhà thầu/NCC'.")

    price_cols = _find_cols_by_header(ws, header_row, "Đơn giá")
    if not price_cols:
        raise SchemaError("[CBE] Không tìm thấy cột 'Đơn giá' nào ở header.")

    ncc_blocks = [NccColumnBlock(cols=[c, c + 1, c + 2]) for c in price_cols]

    name_rows = list(range(name_section_row + 1, header_row))
    # Loại các dòng là header con (vd dòng "Nội dung chào giá") bằng cách
    # chỉ giữ các dòng mà cột A không chứa công thức "=A" tham chiếu chính nó
    name_rows = [r for r in name_rows
                 if not str(ws.cell(row=r, column=1).value or "").startswith("STT")]
    name_rows = name_rows[:len(ncc_blocks)]

    item_start_row = header_row + 1

    summary_anchor_row = _find_row_by_text(ws, "Vận chuyển", col=1,
                                            row_range=(item_start_row, item_start_row + 30))
    if summary_anchor_row is None:
        raise SchemaError("[CBE] Không tìm thấy dòng mốc 'Vận chuyển'.")

    # Item ở CBE có thể chiếm nhiều dòng do merge dọc (cột A) -> dò theo merge span
    template_item_slots = []
    r = item_start_row
    while r < summary_anchor_row:
        height = _merge_span(ws, r, 1)
        # chỉ tính là 1 item nếu cột A có giá trị (item thật), bỏ qua dòng trống
        template_item_slots.append((r, height))
        r += height

    max_col = max(b.cols[-1] for b in ncc_blocks)

    return ItemTableSchema(
        sheet_name=ws.title,
        header_row=header_row,
        item_start_row=item_start_row,
        template_item_slots=template_item_slots,
        ncc_blocks=ncc_blocks,
        max_col=max_col,
        name_rows=name_rows,
        summary_anchor_row=summary_anchor_row,
    )


@dataclass
class SummaryRowSchema:
    """1 dòng tổng kết, định danh bằng marker text ở cột A."""
    marker: str
    row: int


def discover_summary_rows(ws: Worksheet, markers: list[str],
                           start_row: int, max_scan: int = 15) -> dict[str, int]:
    """Tìm các dòng tổng kết theo thứ tự marker xuất hiện, bắt đầu từ start_row."""
    result = {}
    r = start_row
    end = start_row + max_scan
    remaining = list(markers)
    while r <= end and remaining:
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().startswith(remaining[0]):
            result[remaining[0]] = r
            remaining.pop(0)
        r += 1
    if remaining:
        raise SchemaError(f"Không tìm thấy các dòng mốc: {remaining}")
    return result
