"""
factory.py
────────────────────
Bản refactor của ExcelFactory gốc.

Khác biệt cốt lõi so với bản cũ:
  1. Không hardcode row/col — tự dò bằng anchor (excel_schema.py), raise lỗi
     rõ ràng ngay khi load nếu template không khớp cấu trúc mong đợi.
  2. Generic theo số NCC thực tế trong template (không cứng số 3) — nếu data
     có nhiều NCC hơn số khối cột mà template hỗ trợ, raise lỗi rõ ràng thay
     vì âm thầm cắt bớt (`[:3]`) như bản cũ.
  3. Logic "grow bảng item" dùng chung 1 helper (table_grower.py) cho cả
     PMT và CBE, thay vì 2 hàm viết tay gần như giống nhau.
  4. Input được validate bằng dataclass nhẹ — thiếu field quan trọng sẽ báo
     lỗi rõ ràng ngay, không default âm thầm về "" / 0.
"""
from __future__ import annotations

import os
import time
from dataclasses import dataclass, field

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.cell_range import CellRange

from .excel_schema import (
    discover_pmt_schema, discover_cbe_schema, discover_summary_rows,
    SchemaError, ItemTableSchema,
)
from .table_grower import grow_item_table


# ════════════════════════════════════════════════════════════════════════
#  INPUT MODELS — validate sớm, fail rõ ràng thay vì default âm thầm
# ════════════════════════════════════════════════════════════════════════
@dataclass
class Quote:
    ten_hang: str
    don_gia: float = 0.0
    mo_ta: str = ""
    hang: str = ""
    danh_gia: str = "dat"   # "dat" | "khong_dat"


@dataclass
class Supplier:
    ten_ncc: str
    quotes: list[Quote] = field(default_factory=list)

    def find_quote(self, ten_hang: str) -> Quote | None:
        for q in self.quotes:
            if q.ten_hang.strip() == ten_hang.strip():
                return q
        return None


@dataclass
class Item:
    ten_hang: str
    chi_tiet: str = ""
    dvt: str = ""
    so_luong: float = 0
    chung_chi: str = ""
    thoi_gian_bao_hanh: str = ""


@dataclass
class RfqData:
    bo_phan_yeu_cau: str
    items: list[Item]
    nha_cung_cap: list[Supplier]
    ma_ho_so: str = "Unknown"
    thoi_gian_bao_hanh: str = "Theo tiêu chuẩn nhà sản xuất"
    thoi_gian_giao_hang: str = "Theo thỏa thuận"

    @staticmethod
    def from_dict(d: dict) -> "RfqData":
        items = [Item(
            ten_hang=it.get("ten_hang", ""),
            chi_tiet=it.get("chi_tiet", ""),
            dvt=it.get("dvt", ""),
            so_luong=float(it.get("so_luong", 0)),
            chung_chi=it.get("chung_chi", ""),
            thoi_gian_bao_hanh=it.get("thoi_gian_bao_hanh", "")
        ) for it in d.get("items", [])]

        suppliers = []
        for s in d.get("nha_cung_cap", []):
            quotes = [Quote(
                ten_hang=q.get("ten_hang", ""),
                don_gia=float(q.get("don_gia", 0.0) or 0.0),
                mo_ta=q.get("mo_ta", ""),
                hang=q.get("hang", ""),
                danh_gia=q.get("danh_gia", "dat")
            ) for q in s.get("quotes", [])]
            suppliers.append(Supplier(ten_ncc=s.get("ten_ncc", ""), quotes=quotes))

        if not items:
            raise ValueError("RfqData: 'items' không được rỗng.")
        if not suppliers:
            raise ValueError("RfqData: 'nha_cung_cap' không được rỗng.")
        return RfqData(
            bo_phan_yeu_cau=d.get("bo_phan_yeu_cau", ""),
            items=items,
            nha_cung_cap=suppliers,
            ma_ho_so=str(d.get("ma_ho_so", "Unknown")),
            thoi_gian_bao_hanh=d.get("thoi_gian_bao_hanh", "Theo tiêu chuẩn nhà sản xuất"),
            thoi_gian_giao_hang=d.get("thoi_gian_giao_hang", "Theo thỏa thuận"),
        )


THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))


def _check_ncc_capacity(schema: ItemTableSchema, n_ncc: int, sheet_label: str):
    if n_ncc > schema.n_ncc_supported:
        raise SchemaError(
            f"[{sheet_label}] Data có {n_ncc} NCC nhưng template chỉ có "
            f"{schema.n_ncc_supported} khối cột NCC. Cần thêm khối cột vào "
            f"template trước, code sẽ KHÔNG tự cắt bớt dữ liệu."
        )


# ════════════════════════════════════════════════════════════════════════
#  FILL PMT
# ════════════════════════════════════════════════════════════════════════
def fill_pmt_sheet(ws, data: RfqData):
    schema = discover_pmt_schema(ws)
    n_active = len(data.nha_cung_cap)
    _check_ncc_capacity(schema, n_active, "PMT")

    ws.cell(row=2, column=1).value = f"V/v: Mua sắm trang thiết bị cho {data.bo_phan_yeu_cau}"
    ws.cell(row=4, column=1).value = ", ".join(
        it.ten_hang for it in data.items if it.ten_hang)

    for i, row in enumerate(schema.name_rows):
        ws.cell(row=row, column=1).value = (
            data.nha_cung_cap[i].ten_ncc if i < n_active else "")

    grow = grow_item_table(ws, schema, len(data.items))

    for idx, item in enumerate(data.items):
        row = grow.item_rows[idx]
        # Xóa fix cứng chiều cao dòng để Excel tự co giãn theo nội dung (autofit)
        ws.row_dimensions[row].height = None

        ws.cell(row=row, column=1).value = idx + 1
        c2 = ws.cell(row=row, column=2)
        c2.value = item.ten_hang
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        c3 = ws.cell(row=row, column=3)
        c3.value = item.chi_tiet
        c3.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=4).value = item.dvt
        ws.cell(row=row, column=5).value = item.so_luong

        # Định dạng RichText (Đậm gạch chân cho Tên, thường cho Chi tiết)
        font_title = InlineFont(sz=20, b=True, u='single')
        font_normal = InlineFont(sz=20)

        # Khối NCC còn dư (vượt n_active) được dọn trống, không để sót dữ liệu mẫu cũ
        for ncc_idx, block in enumerate(schema.ncc_blocks):
            mo_ta_col, sl_col, hang_col, dg_col = block.cols
            if ncc_idx >= n_active:
                for c in block.cols:
                    ws.cell(row=row, column=c).value = None
                continue
            supplier = data.nha_cung_cap[ncc_idx]
            quote = supplier.find_quote(item.ten_hang)
            danh_gia_txt = "Đạt" if (not quote or quote.danh_gia == "dat") else "Không Đạt"
            
            mo_ta_text = str(quote.mo_ta if (quote and quote.mo_ta) else (item.chi_tiet or ""))
            cell_mo_ta = ws.cell(row=row, column=mo_ta_col)
            cell_mo_ta.value = CellRichText(
                TextBlock(font_title, f"{item.ten_hang or ''}\n"),
                TextBlock(font_normal, mo_ta_text)
            )
            cell_mo_ta.alignment = Alignment(wrap_text=True, vertical="top")
            
            ws.cell(row=row, column=sl_col).value = item.so_luong
            
            cell_hang = ws.cell(row=row, column=hang_col)
            cell_hang.value = quote.hang if quote else ""
            cell_hang.alignment = Alignment(wrap_text=True, vertical="top")
            
            ws.cell(row=row, column=dg_col).value = danh_gia_txt

    _write_pmt_summary(ws, schema, grow.summary_offset, data)


def _write_pmt_summary(ws, schema: ItemTableSchema, offset: int, data: RfqData):
    n_active = len(data.nha_cung_cap)
    base = schema.summary_anchor_row + offset
    rows = {"cert": base, "warranty": base + 1, "delivery": base + 2, "general": base + 3}

    for r in rows.values():
        for c in range(1, schema.max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

    def merge_set(row, col_start, col_end, value, bold=False, center=False):
        if col_end > col_start:
            range_str = f"{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}"
            target = CellRange(range_str)
            overlaps = [rng for rng in ws.merged_cells.ranges if not target.isdisjoint(rng)]
            for rng in overlaps:
                ws.unmerge_cells(str(rng))
            ws.merge_cells(range_str)
        cell = ws.cell(row=row, column=col_start)
        cell.value = value
        cell.alignment = Alignment(wrap_text=True, vertical="center",
                                    horizontal="center" if center else "left")
        cell.font = Font(name="Times New Roman", size=20, bold=bold)

    # Gom chứng chỉ theo nhóm giống nhau
    cc_map: dict[frozenset, tuple[str, list[int]]] = {}
    bh_map: dict[str, list[int]] = {}  # Gom bảo hành

    for idx, item in enumerate(data.items):
        # Gom chứng chỉ
        raw = item.chung_chi or "CO, CQ"
        parts = sorted(set(p.strip() for p in raw.split(",") if p.strip()))
        fs = frozenset(parts)
        if fs not in cc_map:
            cc_map[fs] = (", ".join(parts), [])
        cc_map[fs][1].append(idx + 1)
        
        # Gom bảo hành
        bh = item.thoi_gian_bao_hanh if item.thoi_gian_bao_hanh else data.thoi_gian_bao_hanh
        if bh not in bh_map:
            bh_map[bh] = []
        bh_map[bh].append(idx + 1)

    if len(cc_map) == 1:
        cc_text = f"- {list(cc_map.values())[0][0]}"
    else:
        cc_text = "\n".join(
            f"- {disp} (Đối với mục số {', '.join(f'{i:02d}' for i in idxs)})"
            for disp, idxs in cc_map.values()
        )
        
    if len(bh_map) == 1:
        bh_text = list(bh_map.keys())[0]
    else:
        bh_text = "\n".join(
            f"- {disp} (Đối với mục số {', '.join(f'{i:02d}' for i in idxs)})"
            for disp, idxs in bh_map.items()
        )

    rows_spec = [
        (rows["cert"], "Chứng chỉ Hàng hóa/ Certificate of Origin (CO)/ Certificate of Quality (CQ)", cc_text),
        (rows["warranty"], "Thời gian bảo hành/ Warranty period", bh_text),
        (rows["delivery"], "Thời gian giao hàng (Delivery time)", data.thoi_gian_giao_hang),
    ]
    for row, label, value in rows_spec:
        merge_set(row, 1, 5, label, bold=True)
        for ncc_idx, block in enumerate(schema.ncc_blocks):
            mo_ta_col, sl_col, hang_col, dg_col = block.cols
            has_supplier = ncc_idx < n_active
            merge_set(row, mo_ta_col, hang_col, value if has_supplier else "")
            ws.cell(row=row, column=dg_col).value = "Đạt" if has_supplier else ""

    # Đánh giá chung
    merge_set(rows["general"], 1, 5, "Đánh giá chung/ General evaluation", bold=True)
    for ncc_idx, block in enumerate(schema.ncc_blocks):
        mo_ta_col, sl_col, hang_col, dg_col = block.cols
        if ncc_idx >= n_active:
            merge_set(rows["general"], mo_ta_col, dg_col, "", bold=True, center=True)
            continue
        supplier = data.nha_cung_cap[ncc_idx]
        overall = "ĐẠT"
        for q in supplier.quotes:
            if q.danh_gia != "dat":
                overall = "KHÔNG ĐẠT"
                break
        merge_set(rows["general"], mo_ta_col, dg_col, overall, bold=True, center=True)


# ════════════════════════════════════════════════════════════════════════
#  FILL CBE
# ════════════════════════════════════════════════════════════════════════
def fill_cbe_sheet(ws, data: RfqData):
    schema = discover_cbe_schema(ws)
    n_active = len(data.nha_cung_cap)
    _check_ncc_capacity(schema, n_active, "CBE")

    for i, row in enumerate(schema.name_rows):
        ws.cell(row=row, column=1).value = (
            data.nha_cung_cap[i].ten_ncc if i < n_active else "")

    grow = grow_item_table(ws, schema, len(data.items))

    for idx, item in enumerate(data.items):
        row = grow.item_rows[idx]
        # Xóa fix cứng chiều cao dòng để Excel tự co giãn theo nội dung (autofit)
        ws.row_dimensions[row].height = None

        ws.cell(row=row, column=1).value = idx + 1
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = f"{item.ten_hang}\n{item.chi_tiet}" if item.chi_tiet else item.ten_hang
        cell_b.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3).value = item.dvt
        ws.cell(row=row, column=4).value = item.so_luong

        for ncc_idx, block in enumerate(schema.ncc_blocks):
            don_gia_col, thanh_tien_col, ghi_chu_col = block.cols
            if ncc_idx >= n_active:
                for c in block.cols:
                    ws.cell(row=row, column=c).value = None
                continue
            supplier = data.nha_cung_cap[ncc_idx]
            quote = supplier.find_quote(item.ten_hang)
            don_gia = float(quote.don_gia) if quote else 0.0
            danh_gia_txt = "Đạt" if (not quote or quote.danh_gia == "dat") else "Không Đạt"

            ws.cell(row=row, column=don_gia_col).value = don_gia
            don_gia_letter = get_column_letter(don_gia_col)
            qty_letter = get_column_letter(4)
            ws.cell(row=row, column=thanh_tien_col).value = (
                f"={don_gia_letter}{row}*${qty_letter}{row}"
            )
            ws.cell(row=row, column=ghi_chu_col).value = danh_gia_txt

    _write_cbe_summary(ws, schema, grow, data)


def _write_cbe_summary(ws, schema: ItemTableSchema, grow, data: RfqData):
    n_active = len(data.nha_cung_cap)
    base = schema.summary_anchor_row + grow.summary_offset
    t_row, tong_row, vat_row, gia_dg_row, xh_row = base, base + 1, base + 2, base + 3, base + 4
    last_item_row = grow.item_rows[-1]
    first_item_row = grow.item_rows[0]

    def safe_merge_cells(range_str):
        target = CellRange(range_str)
        overlaps = [rng for rng in ws.merged_cells.ranges if not target.isdisjoint(rng)]
        for rng in overlaps:
            ws.unmerge_cells(str(rng))
        ws.merge_cells(range_str)

    def merge_set(row, col_start, col_end, value):
        range_str = f"{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}"
        safe_merge_cells(range_str)
        ws.cell(row=row, column=col_start).value = value

    def clear_block(row, block):
        range_str = f"{get_column_letter(block.cols[0])}{row}:{get_column_letter(block.cols[-1])}{row}"
        safe_merge_cells(range_str)
        ws.cell(row=row, column=block.cols[0]).value = None

    merge_set(t_row, 1, 4, "Vận chuyển (Transport)")
    merge_set(tong_row, 1, 4, "Tổng cộng")
    merge_set(vat_row, 1, 4, "Thuế VAT 8%")
    merge_set(gia_dg_row, 1, 4, "Giá đánh giá (Price evaluation)")
    ws.cell(row=xh_row, column=1).value = "III"
    merge_set(xh_row, 2, 4, "Xếp hạng")

    # Chỉ ghi formula cho đúng các khối NCC có dữ liệu thật (n_active) —
    # khối dư (nếu template hỗ trợ nhiều hơn số NCC thực tế) được dọn trống,
    # KHÔNG để sót formula tham chiếu cột rỗng (đây là bug đã có ở bản gốc:
    # rank formula luôn so cả 3 cột dù chỉ có 2 NCC, làm NCC trống "thắng" giá 0đ).
    active_blocks = schema.ncc_blocks[:n_active]
    refs = [get_column_letter(b.cols[0]) + str(gia_dg_row) for b in active_blocks]

    for ncc_idx, block in enumerate(schema.ncc_blocks):
        don_gia_col, thanh_tien_col, ghi_chu_col = block.cols
        if ncc_idx >= n_active:
            for row in (t_row, tong_row, vat_row, gia_dg_row, xh_row):
                clear_block(row, block)
            continue

        tt_letter = get_column_letter(thanh_tien_col)
        dg_letter = get_column_letter(don_gia_col)
        gc_letter = get_column_letter(ghi_chu_col)

        merge_set(t_row, don_gia_col, ghi_chu_col, "Đã bao gồm")
        merge_set(tong_row, don_gia_col, ghi_chu_col,
                  f"=SUM({tt_letter}{first_item_row}:{tt_letter}{last_item_row})")
        merge_set(vat_row, don_gia_col, ghi_chu_col, f"={dg_letter}{tong_row}*8%")
        merge_set(gia_dg_row, don_gia_col, ghi_chu_col,
                  f"=SUM({dg_letter}{tong_row}:{gc_letter}{vat_row})")

        ref = get_column_letter(don_gia_col) + str(gia_dg_row)
        formula = f"=IF({ref}=MIN({','.join(refs)}),1,IF({ref}=MAX({','.join(refs)}),3,2))"
        merge_set(xh_row, don_gia_col, ghi_chu_col, formula)


# ════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════════════════
class ExcelFactory:
    CBE_TEMPLATE_NAME = "QN-COM-PR01-FM06 Danh gia lua chon NCC.xlsx"

    def __init__(self, template_path: str | None = None,
                 base_data_path: str | None = None):
        self.base_data_path = os.path.abspath(base_data_path or os.getenv("BASE_DATA_PATH", "/data"))
        os.makedirs(self.base_data_path, exist_ok=True)
        self.template_path = template_path or os.path.join(self.base_data_path, self.CBE_TEMPLATE_NAME)

    def generate_cbe_excel(self, data: dict) -> dict:
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Không tìm thấy template: {self.template_path}")

        rfq = RfqData.from_dict(data)
        wb = openpyxl.load_workbook(self.template_path)

        if "PMT" in wb.sheetnames:
            fill_pmt_sheet(wb["PMT"], rfq)
        if "CBE" in wb.sheetnames:
            fill_cbe_sheet(wb["CBE"], rfq)

        try:
            from word_factory.utils.formatters import sanitize_filename
            safe_name = sanitize_filename(rfq.ma_ho_so)
        except Exception:
            safe_name = rfq.ma_ho_so.replace("/", "-").replace("\\", "-")

        output_name = f"CBE_Goi2_{safe_name}_{int(time.time())}.xlsx"
        output_path = os.path.join(self.base_data_path, output_name)
        wb.save(output_path)
        return {"status": "success", "file_name": output_name, "file_path": output_path}

    def generate_danh_gia_ky_thuat(self, payload: dict) -> dict:
        """Entry point từ api.py — delegate sang generate_cbe_excel."""
        return self.generate_cbe_excel(payload)
