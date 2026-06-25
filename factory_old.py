import openpyxl
import os
import time
from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, PatternFill

class ExcelFactory:
    # ── Tên file template ────────────────────────────────────────────────
    CBE_TEMPLATE_NAME = "QN-COM-PR01-FM06 Danh gia lua chon NCC.xlsx"

    # ── Sheet PMT: dòng bắt đầu ghi items ───────────────────────────────
    PMT_ITEM_START_ROW = 11   # Dựa theo điều tra thực tế: items bắt đầu từ row 11

    # ── Sheet CBE: dòng bắt đầu ghi items ───────────────────────────────
    CBE_ITEM_START_ROW = 20   # Dựa theo điều tra thực tế: items bắt đầu từ row 20

    # ── CBE: Mapping cột cho từng NCC ───────────────────────────────────
    # Mỗi NCC chiếm 3 cột: (don_gia_col, thanh_tien_col, danh_gia_col)
    # Thành tiền = cột tính = =don_gia * so_luong
    # Danh giá   = "Đạt" / "Không Đạt"
    CBE_NCC_COL_MAP = {
        0: (5,  6,  7),   # NCC1: E=đơn giá, F=thành tiền, G=ghi chú/đánh giá
        1: (8,  9,  10),  # NCC2: H=đơn giá, I=thành tiền, J=ghi chú/đánh giá
        2: (11, 12, 13),  # NCC3: K=đơn giá, L=thành tiền, M=ghi chú/đánh giá
    }

    # ── PMT: Mapping cột cho từng NCC ───────────────────────────────────
    # Mỗi NCC chiếm 4 cột: (mo_ta_col, so_luong_col, hang_col, danh_gia_col)
    # Cột B=tên hàng, C=chi tiết, D=dvt, E=số lượng (cột PTSC yêu cầu)
    PMT_NCC_COL_MAP = {
        0: (6,  7,  8,  9),   # NCC1: F=mô tả chào, G=SL, H=hãng/xuất xứ, I=đánh giá
        1: (10, 11, 12, 13),  # NCC2: J=mô tả chào, K=SL, L=hãng/xuất xứ, M=đánh giá
        2: (14, 15, 16, 17),  # NCC3: N=mô tả chào, O=SL, P=hãng/xuất xứ, Q=đánh giá
    }

    def __init__(self):
        self.base_data_path = os.path.abspath(os.getenv("BASE_DATA_PATH", "/data"))
        os.makedirs(self.base_data_path, exist_ok=True)

        # Tìm template: BASE_DATA_PATH trước, fallback sang ../data
        script_dir  = os.path.dirname(os.path.abspath(__file__))
        fallback    = os.path.join(script_dir, "..", "data")
        candidate1  = os.path.join(self.base_data_path, self.CBE_TEMPLATE_NAME)
        candidate2  = os.path.abspath(os.path.join(fallback, self.CBE_TEMPLATE_NAME))

        if os.path.exists(candidate1):
            self.template_path = candidate1
        elif os.path.exists(candidate2):
            self.template_path = candidate2
        else:
            self.template_path = candidate1  # sẽ raise lỗi ở generate nếu thiếu

    # ════════════════════════════════════════════════════════════════════
    #  HELPER: Copy toàn bộ style từ ô nguồn sang ô đích
    # ════════════════════════════════════════════════════════════════════
    @staticmethod
    def _copy_cell_style(src, dst):
        """Copy font, border, fill, alignment, number_format."""
        if src.has_style:
            dst.font        = copy(src.font)
            dst.border      = copy(src.border)
            dst.fill        = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection  = copy(src.protection)
            dst.alignment   = copy(src.alignment)

    # ════════════════════════════════════════════════════════════════════
    #  HELPER: Copy style toàn bộ một dòng từ source_row sang target_row
    # ════════════════════════════════════════════════════════════════════
    @staticmethod
    def _copy_row_style(ws, source_row: int, target_row: int, max_col: int):
        for col in range(1, max_col + 1):
            src = ws.cell(row=source_row, column=col)
            dst = ws.cell(row=target_row, column=col)
            if src.has_style:
                dst.font        = copy(src.font)
                dst.border      = copy(src.border)
                dst.fill        = copy(src.fill)
                dst.number_format = src.number_format
                dst.protection  = copy(src.protection)
                dst.alignment   = copy(src.alignment)

    # ════════════════════════════════════════════════════════════════════
    #  HELPER: Unmerge tất cả merged cells trong một vùng dòng
    # ════════════════════════════════════════════════════════════════════
    @staticmethod
    def _unmerge_rows(ws, from_row: int, to_row: int):
        to_remove = [
            str(r) for r in ws.merged_cells.ranges
            if r.min_row >= from_row and r.max_row <= to_row
        ]
        for rng in to_remove:
            ws.unmerge_cells(rng)

    # ════════════════════════════════════════════════════════════════════
    #  HELPER: Xóa giá trị (giữ style) trong vùng hình chữ nhật
    # ════════════════════════════════════════════════════════════════════
    @staticmethod
    def _clear_values(ws, from_row: int, to_row: int, max_col: int):
        for r in range(from_row, to_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).value = None

    # ════════════════════════════════════════════════════════════════════
    #  HELPER: Tìm báo giá của NCC cho một tên hàng cụ thể
    # ════════════════════════════════════════════════════════════════════
    @staticmethod
    def _find_quote(ncc: dict, ten_hang: str) -> dict | None:
        for q in ncc.get("quotes", []):
            if q.get("ten_hang", "").strip() == ten_hang.strip():
                return q
        return None

    # ════════════════════════════════════════════════════════════════════
    #  MAIN ENTRY POINT
    # ════════════════════════════════════════════════════════════════════
    def generate_cbe_excel(self, data: dict) -> dict:
        """
        Đúc file Excel Gói 2 từ template.
        Điền đầy đủ vào cả 2 sheet: PMT (kỹ thuật) và CBE (thương mại).

        data = {
          bo_phan_yeu_cau: str,
          items: [{ten_hang, chi_tiet, dvt, so_luong}, ...],
          nha_cung_cap: [{ten_ncc, quotes: [{ten_hang, don_gia, danh_gia}]}, ...]
        }
        """
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(
                f"[Goi2] Khong tim thay template: {self.template_path}\n"
                f"       Dat file '{self.CBE_TEMPLATE_NAME}' vao thu muc data."
            )

        print(f"[Goi2] Mo template: {self.template_path}")
        wb = openpyxl.load_workbook(self.template_path)

        # ── Bước 1: Fill sheet PMT (Đánh giá Kỹ thuật) ──────────────────
        self._fill_pmt_sheet(wb, data)

        # ── Bước 2: Fill sheet CBE (Đánh giá Thương mại) ─────────────────
        self._fill_cbe_sheet(wb, data)

        # ── Bước 3: Lưu file output ───────────────────────────────────────
        ma_ho_so = str(data.get("ma_ho_so", "Unknown"))
        try:
            from word_factory.utils.formatters import sanitize_filename
            safe_name = sanitize_filename(ma_ho_so)
        except Exception:
            safe_name = ma_ho_so.replace("/", "-").replace("\\", "-")

        output_name = f"CBE_Goi2_{safe_name}_{int(time.time())}.xlsx"
        output_path = os.path.join(self.base_data_path, output_name)
        wb.save(output_path)
        print(f"[Goi2] Luu thanh cong: {output_path}")
        return {"status": "success", "file_name": output_name}

    # ════════════════════════════════════════════════════════════════════
    #  FILL SHEET PMT — Bảng Đánh Giá Kỹ Thuật (17 cột A-Q)
    # ════════════════════════════════════════════════════════════════════
    def _fill_pmt_sheet(self, wb: openpyxl.Workbook, data: dict):
        """
        Điền vào sheet PMT theo cấu trúc thực tế đã điều tra:
          - A2        : Trích yếu V/v (merge A2:O2)
          - A4        : Tên các hàng hóa (cách nhau dấu phẩy)
          - A6/A7/A8  : Tên NCC1/NCC2/NCC3
          - Row 11+   : Từng item, mỗi item 1 dòng, 17 cột (A-Q)
        """
        if "PMT" not in wb.sheetnames:
            print("[Goi2][PMT] Khong co sheet PMT, bo qua.")
            return

        ws          = wb["PMT"]
        items       = data.get("items", [])
        nha_cung_cap = data.get("nha_cung_cap", [])
        bo_phan     = data.get("bo_phan_yeu_cau", "")

        # ── 1. Ghi thông tin chung ──────────────────────────────────────
        # A2 (merge A2:O2): Trích yếu
        ws["A2"] = f"V/v: Mua sắm trang thiết bị cho {bo_phan}"

        # A4: Danh sách tên hàng hóa
        ten_hang_list = ", ".join(
            item.get("ten_hang", "") for item in items if item.get("ten_hang")
        )
        ws["A4"] = ten_hang_list

        # A6, A7, A8: Tên NCC (tối đa 3, để trống nếu thiếu)
        for i, addr in enumerate(["A6", "A7", "A8"]):
            ws[addr] = nha_cung_cap[i].get("ten_ncc", "") if i < len(nha_cung_cap) else ""

        print(f"[Goi2][PMT] Da ghi header: bo_phan='{bo_phan}', "
              f"{len(nha_cung_cap)} NCC, {len(items)} items")

        # ── 2. Chuẩn bị vùng items ──────────────────────────────────────
        start = self.PMT_ITEM_START_ROW
        # Vùng xóa: từ start đến start + max(items)+5 để sạch data cũ
        end_clear = start + max(len(items), 2) + 3
        MAX_COL_PMT = 17  # A-Q

        # Unmerge các ô trong vùng items (nếu có)
        self._unmerge_rows(ws, start, end_clear)
        # Xóa giá trị cũ (giữ style)
        self._clear_values(ws, start, end_clear, MAX_COL_PMT)

        # ── 3. Ghi từng item ────────────────────────────────────────────
        for idx, item in enumerate(items):
            row      = start + idx
            ten_hang = str(item.get("ten_hang", ""))
            chi_tiet = str(item.get("chi_tiet", ""))
            dvt      = str(item.get("dvt", ""))
            so_luong = item.get("so_luong", 0)

            # Copy style từ dòng mẫu (row 11 = dòng item đầu tiên trong template)
            if idx > 0:
                self._copy_row_style(ws, start, row, MAX_COL_PMT)

            # Cột A: STT
            ws.cell(row=row, column=1).value = idx + 1

            # Cột B: Tên hàng (ngắn gọn)
            ws.cell(row=row, column=2).value = ten_hang

            # Cột C: Chi tiết kỹ thuật PTSC yêu cầu (wrap_text)
            cell_c = ws.cell(row=row, column=3)
            cell_c.value = chi_tiet
            cell_c.alignment = Alignment(wrap_text=True, vertical="top")

            # Cột D: DVT
            ws.cell(row=row, column=4).value = dvt

            # Cột E: Số lượng PTSC yêu cầu
            ws.cell(row=row, column=5).value = so_luong

            # ── Điền dữ liệu từng NCC vào khối cột tương ứng ────────
            for ncc_idx, ncc in enumerate(nha_cung_cap[:3]):
                col_mo_ta, col_so_luong, col_hang, col_danh_gia = \
                    self.PMT_NCC_COL_MAP[ncc_idx]

                quote       = self._find_quote(ncc, ten_hang)
                danh_gia_raw = quote.get("danh_gia", "dat") if quote else "dat"
                danh_gia_txt = "Đạt" if danh_gia_raw == "dat" else "Không Đạt"

                # Cột F/J/N: Mô tả hàng hóa NCC chào thầu (Option A: placeholder)
                cell_mo_ta = ws.cell(row=row, column=col_mo_ta)
                cell_mo_ta.value = "-"
                cell_mo_ta.alignment = Alignment(wrap_text=True, vertical="top")

                # Cột G/K/O: Số lượng NCC (copy từ cột E)
                ws.cell(row=row, column=col_so_luong).value = f"=E{row}"

                # Cột H/L/P: Hãng/Xuất xứ (Option A: placeholder)
                ws.cell(row=row, column=col_hang).value = "-"

                # Cột I/M/Q: Kết quả đánh giá
                ws.cell(row=row, column=col_danh_gia).value = danh_gia_txt

        print(f"[Goi2][PMT] Da ghi {len(items)} items tu dong {start}")

    # ════════════════════════════════════════════════════════════════════
    #  FILL SHEET CBE — Bảng Đánh Giá Thương Mại (13 cột A-M)
    # ════════════════════════════════════════════════════════════════════
    def _fill_cbe_sheet(self, wb: openpyxl.Workbook, data: dict):
        """
        Điền vào sheet CBE theo cấu trúc thực tế đã điều tra:
          - A7/A8/A9  : Tên NCC1/NCC2/NCC3 (độc lập, không link PMT)
          - Row 20+   : Từng item, mỗi item 1 dòng, 13 cột (A-M)
          - Sau items : Summary rows trượt động (Transport, Tổng, VAT, GiáĐG, XếpHạng)
        """
        if "CBE" not in wb.sheetnames:
            raise ValueError("[Goi2][CBE] Khong tim thay sheet 'CBE' trong template!")

        ws           = wb["CBE"]
        items        = data.get("items", [])
        nha_cung_cap = data.get("nha_cung_cap", [])
        MAX_COL_CBE  = 13  # A-M

        # ── 1. Ghi tên NCC vào A7, A8, A9 ─────────────────────────────
        # (CBE dùng =A7, =A8, =A9 làm header bảng — KHÔNG link từ PMT)
        for i, addr in enumerate(["A7", "A8", "A9"]):
            ws[addr] = nha_cung_cap[i].get("ten_ncc", "") if i < len(nha_cung_cap) else ""

        print(f"[Goi2][CBE] Da ghi ten NCC: {[n.get('ten_ncc','') for n in nha_cung_cap[:3]]}")

        # ── 2. Chuẩn bị vùng items ─────────────────────────────────────
        start     = self.CBE_ITEM_START_ROW
        # Buffer xóa đủ rộng: items + 10 dòng summary buffer
        end_clear = start + max(len(items), 2) + 10

        # Unmerge toàn bộ vùng sẽ ghi (template có merge 2 dòng cho item 1)
        self._unmerge_rows(ws, start, end_clear)
        # Xóa data cũ (giữ style)
        self._clear_values(ws, start, end_clear, MAX_COL_CBE)

        # ── 3. Ghi từng item (mỗi item = 1 dòng đơn, wrap_text) ────────
        for idx, item in enumerate(items):
            row      = start + idx
            ten_hang = str(item.get("ten_hang", ""))
            chi_tiet = str(item.get("chi_tiet", ""))
            dvt      = str(item.get("dvt", ""))
            so_luong = item.get("so_luong", 0)

            # Copy style từ dòng mẫu row 20 cho các dòng thêm mới
            if idx > 0:
                self._copy_row_style(ws, start, row, MAX_COL_CBE)

            # Cột A: STT
            ws.cell(row=row, column=1).value = idx + 1

            # Cột B: Tên hàng + Chi tiết kỹ thuật (xuống dòng, wrap_text)
            cell_b = ws.cell(row=row, column=2)
            cell_b.value = f"{ten_hang}\n{chi_tiet}" if chi_tiet else ten_hang
            cell_b.alignment = Alignment(wrap_text=True, vertical="top")

            # Cột C: DVT
            ws.cell(row=row, column=3).value = dvt

            # Cột D: Số lượng
            ws.cell(row=row, column=4).value = so_luong

            # ── Điền dữ liệu từng NCC vào đúng khối cột ──────────────
            for ncc_idx, ncc in enumerate(nha_cung_cap[:3]):
                col_dg, col_tt, col_kq = self.CBE_NCC_COL_MAP[ncc_idx]

                quote        = self._find_quote(ncc, ten_hang)
                don_gia      = float(quote.get("don_gia", 0)) if quote else 0.0
                danh_gia_raw = quote.get("danh_gia", "dat") if quote else "dat"
                danh_gia_txt = "Đạt" if danh_gia_raw == "dat" else "Không Đạt"

                don_gia_letter = get_column_letter(col_dg)

                # Đơn giá
                ws.cell(row=row, column=col_dg).value = don_gia

                # Thành tiền = đơn giá × số lượng (công thức Excel)
                ws.cell(row=row, column=col_tt).value = (
                    f"={don_gia_letter}{row}*$D{row}"
                )

                # Ghi chú / Kết quả đánh giá
                ws.cell(row=row, column=col_kq).value = danh_gia_txt

        print(f"[Goi2][CBE] Da ghi {len(items)} items tu dong {start}")

        # ── 4. Ghi summary rows trượt động ──────────────────────────────
        self._write_cbe_summary_rows(ws, start, len(items))

    # ════════════════════════════════════════════════════════════════════
    #  WRITE CBE SUMMARY ROWS — Transport, Tổng, VAT, GiáĐG, Xếp hạng
    # ════════════════════════════════════════════════════════════════════
    def _write_cbe_summary_rows(self, ws, item_start: int, item_count: int):
        """
        Ghi và re-merge đúng các dòng tổng kết sau vùng items.
        Vị trí trượt theo item_count:
          transport_row = item_start + item_count
          tong_row      = transport_row + 1
          vat_row       = tong_row + 1
          gia_dg_row    = vat_row + 1
          xep_hang_row  = gia_dg_row + 1
        """
        last_item  = item_start + item_count - 1
        t_row      = last_item + 1   # Transport
        tong_row   = last_item + 2   # Tổng cộng
        vat_row    = last_item + 3   # VAT 8%
        gia_dg_row = last_item + 4   # Giá đánh giá
        xh_row     = last_item + 5   # Xếp hạng

        # Helper: merge một dải ô theo pattern cố định của CBE
        # Pattern: (A:D), (E:G), (H:J), (K:M) — mỗi dòng summary đều dùng pattern này
        def _merge_and_set(row, col_start, col_end, value):
            """Merge từ col_start đến col_end ở row, ghi value vào master cell."""
            rng = f"{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}"
            ws.merge_cells(rng)
            ws.cell(row=row, column=col_start).value = value

        # ── Transport ────────────────────────────────────────────────────
        _merge_and_set(t_row, 1,  4,  "Vận chuyển (Transport)")
        _merge_and_set(t_row, 5,  7,  "Đã bao gồm")
        _merge_and_set(t_row, 8,  10, "Đã bao gồm")
        _merge_and_set(t_row, 11, 13, "Đã bao gồm")

        # ── Tổng cộng ────────────────────────────────────────────────────
        # SUM theo cột Thành tiền: F (col 6) cho NCC1, I (col 9) cho NCC2, L (col 12) cho NCC3
        # Kết quả ghi vào master cell của block merge: E (col 5), H (col 8), K (col 11)
        _merge_and_set(tong_row, 1, 4, "Tổng cộng")
        _merge_and_set(tong_row, 5,  7,
                       f"=SUM(F{item_start}:F{last_item})")   # SUM cột F (Thành tiền NCC1)
        _merge_and_set(tong_row, 8,  10,
                       f"=SUM(I{item_start}:I{last_item})")   # SUM cột I (Thành tiền NCC2)
        _merge_and_set(tong_row, 11, 13,
                       f"=SUM(L{item_start}:L{last_item})")   # SUM cột L (Thành tiền NCC3)

        # ── VAT 8% ───────────────────────────────────────────────────────
        # Tham chiếu master cell của Tổng cộng: E{tong}, H{tong}, K{tong}
        _merge_and_set(vat_row, 1, 4, "Thuế VAT 8%")
        _merge_and_set(vat_row, 5,  7,  f"=E{tong_row}*8%")
        _merge_and_set(vat_row, 8,  10, f"=H{tong_row}*8%")
        _merge_and_set(vat_row, 11, 13, f"=K{tong_row}*8%")

        # ── Giá đánh giá ─────────────────────────────────────────────────
        # =SUM(E{tong}:G{vat}) → tổng toàn bộ block 3 cột × 2 dòng của mỗi NCC
        # (giống hệt công thức gốc trong template: =SUM(E24:G25))
        _merge_and_set(gia_dg_row, 1, 4, "Giá đánh giá (Price evaluation)")
        _merge_and_set(gia_dg_row, 5,  7,
                       f"=SUM(E{tong_row}:G{vat_row})")   # NCC1 block
        _merge_and_set(gia_dg_row, 8,  10,
                       f"=SUM(H{tong_row}:J{vat_row})")   # NCC2 block
        _merge_and_set(gia_dg_row, 11, 13,
                       f"=SUM(K{tong_row}:M{vat_row})")   # NCC3 block

        # ── Xếp hạng ─────────────────────────────────────────────────────
        # Xếp hạng 1/2/3 dựa trên Giá đánh giá (nhỏ nhất = hạng 1)
        rank_formula = (
            "=IF({ref}=MIN($E${r},$H${r},$K${r}),1,"
            "IF({ref}=MAX($E${r},$H${r},$K${r}),3,2))"
        )
        ws.cell(row=xh_row, column=1).value = "III"
        _merge_and_set(xh_row, 2, 4, "Xếp hạng")
        _merge_and_set(xh_row, 5,  7,
                       rank_formula.format(ref=f"E{gia_dg_row}", r=gia_dg_row))
        _merge_and_set(xh_row, 8,  10,
                       rank_formula.format(ref=f"H{gia_dg_row}", r=gia_dg_row))
        _merge_and_set(xh_row, 11, 13,
                       rank_formula.format(ref=f"K{gia_dg_row}", r=gia_dg_row))

        print(
            f"[Goi2][CBE] Summary rows: "
            f"transport={t_row}, tong={tong_row}, vat={vat_row}, "
            f"gia_dg={gia_dg_row}, xep_hang={xh_row}"
        )

    # ════════════════════════════════════════════════════════════════════
    #  LEGACY ENTRY POINT (backward compat với api.py)
    # ════════════════════════════════════════════════════════════════════
    def generate_danh_gia_ky_thuat(self, payload: dict) -> dict:
        """Entry point từ api.py — delegate sang generate_cbe_excel."""
        return self.generate_cbe_excel(payload)
